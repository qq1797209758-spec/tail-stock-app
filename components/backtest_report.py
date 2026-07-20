"""历史回测 Excel 报告导出。"""

from __future__ import annotations

from io import BytesIO

import pandas as pd

from strategy.backtest import BacktestResult


def build_backtest_excel(result: BacktestResult) -> bytes:
    """生成回测汇总、日收益、明细、失败记录和参数快照。"""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    summary = pd.DataFrame(
        [{"指标": key, "数值": value} for key, value in result.summary.items()]
    )
    parameters = pd.DataFrame(
        [{"参数": key, "当前值": value} for key, value in result.parameters.items()]
    )
    sheets = {
        "回测汇总": summary,
        "每日收益": result.daily.copy(),
        "交易明细": result.details.copy(),
        "失败与缺失": result.failures.copy(),
        "参数快照": parameters,
    }
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, data in sheets.items():
            frame = data.astype(object).where(pd.notna(data), None)
            if frame.empty and not len(frame.columns):
                frame = pd.DataFrame({"说明": ["本次回测无记录"]})
            frame.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
            sheet = writer.book[sheet_name]
            column_count = max(1, len(frame.columns))
            sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=column_count)
            title = sheet.cell(1, 1, f"A股尾盘策略历史回测 · {sheet_name}")
            title.font = Font(bold=True, color="FFFFFF", size=14)
            title.fill = PatternFill("solid", fgColor="102A43")
            title.alignment = Alignment(horizontal="left")
            for cell in sheet[3]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="176B87")
                cell.alignment = Alignment(horizontal="center")
            sheet.freeze_panes = "A4"
            sheet.auto_filter.ref = f"A3:{get_column_letter(column_count)}{sheet.max_row}"
            for index, column in enumerate(frame.columns, start=1):
                values = [str(column), *[str(value) for value in frame[column].dropna().head(200)]]
                sheet.column_dimensions[get_column_letter(index)].width = min(
                    45, max(11, max(map(len, values), default=10) + 2)
                )
                for cell in sheet[get_column_letter(index)][3:]:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                if column in {"数据完整性", "胜率", "最大回撤"}:
                    for cell in sheet[get_column_letter(index)][3:]:
                        cell.number_format = "0.0%"
                elif "收益率" in str(column):
                    for cell in sheet[get_column_letter(index)][3:]:
                        cell.number_format = "0.00"
                elif column == "当日总市值":
                    for cell in sheet[get_column_letter(index)][3:]:
                        cell.number_format = "#,##0"
            if sheet_name == "回测汇总":
                sheet.column_dimensions["A"].width = 38
                sheet.column_dimensions["B"].width = 20
            elif sheet_name == "参数快照":
                sheet.column_dimensions["A"].width = 28
                sheet.column_dimensions["B"].width = 72
            elif sheet_name == "失败与缺失":
                sheet.column_dimensions["D"].width = 72
            sheet.sheet_view.showGridLines = False
            sheet.cell(
                sheet.max_row + 2,
                1,
                "历史回测结果不代表未来表现；本报告仅用于公开数据筛选和策略研究，不构成投资建议。",
            )
    return output.getvalue()
