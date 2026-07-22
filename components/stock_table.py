"""响应式股票结果组件与 Excel 导出。"""

from datetime import datetime
from html import escape
from io import BytesIO

import numpy as np
import pandas as pd
import streamlit as st

from config import MARKET_CAP_UNIT


DISPLAY_COLUMNS = [
    "排名",
    "代码",
    "名称",
    "最新价",
    "涨跌幅",
    "换手率",
    "量比",
    "总市值",
    "最近涨停日期",
    "尾盘结构状态",
    "VWAP状态",
    "高于VWAP占比",
    "尾盘最大回撤",
    "最后10分钟涨跌幅",
    "连续走弱状态",
    "尾盘成交量状态",
    "尾盘结构评分",
    "分钟数据源",
    "分钟K线条数",
    "接口错误原因",
    "淘汰原因",
    "数据完整性",
    "综合得分",
    "观察标记",
    "资金表现得分",
    "板块强度得分",
    "技术形态得分",
    "尾盘结构得分",
    "市场情绪得分",
    "主力资金净流入",
    "所属行业",
    "近5日板块强度",
    "主力净流入占比",
    "缺失字段",
    "数据完整度",
    "当前行情数据源",
    "入选类型",
    "数据更新时间",
    "当前北京时间",
    "评分依据",
    "入选原因",
    "主要风险",
    "次日观察条件",
]


def _display_data(data: pd.DataFrame) -> pd.DataFrame:
    available_columns = [column for column in DISPLAY_COLUMNS if column in data.columns]
    result = data.loc[:, available_columns].copy()
    result.rename(
        columns={
            "涨跌幅": "涨跌幅 (%)",
            "最新价": "当前价",
            "换手率": "换手率 (%)",
            "总市值": "总市值 (亿元)",
            "综合得分": "综合评分",
            "资金表现得分": "资金分",
            "板块强度得分": "板块分",
            "技术形态得分": "技术分",
            "尾盘结构得分": "尾盘分",
            "市场情绪得分": "市场情绪分",
        },
        inplace=True,
    )
    result["总市值 (亿元)"] = result["总市值 (亿元)"] / MARKET_CAP_UNIT
    return result


def _format_number(value: object, digits: int = 2) -> str:
    if pd.isna(value) or not np.isfinite(float(value)):
        return "--"
    return f"{float(value):.{digits}f}"


def _format_display_value(column: str, value: object) -> str:
    if pd.isna(value):
        return "--"
    numeric_columns = {
        "最新价",
        "当前价",
        "涨跌幅 (%)",
        "换手率 (%)",
        "量比",
        "总市值 (亿元)",
        "综合评分",
        "资金分",
        "板块分",
        "技术分",
        "尾盘分",
        "市场情绪分",
        "近5日板块强度",
        "主力净流入占比",
        "主力资金净流入",
    }
    if column == "尾盘结构评分":
        return _format_number(value, 0)
    if column in {
        "高于VWAP占比", "尾盘最大回撤", "最后10分钟涨跌幅", "数据完整性",
        "数据完整度",
    }:
        if not np.isfinite(float(value)):
            return "数据不足"
        return f"{float(value):.1%}"
    if column in numeric_columns:
        return _format_number(value)
    return escape(str(value))


def render_stock_results(data: pd.DataFrame) -> None:
    """桌面显示表格，窄屏显示卡片，数据均来自真实筛选结果。"""
    display = _display_data(data)
    headers = "".join(f"<th>{escape(str(column))}</th>" for column in display.columns)
    rows = []
    cards = []

    for _, row in display.iterrows():
        values = [
            _format_display_value(column, row[column]) for column in display.columns
        ]
        rows.append("<tr>" + "".join(f"<td>{value}</td>" for value in values) + "</tr>")
        code = _format_display_value("代码", row.get("代码"))
        name = _format_display_value("名称", row.get("名称"))
        card_fields = "".join(
            f"<div><small>{escape(str(column))}</small><b>{_format_display_value(column, row[column])}</b></div>"
            for column in display.columns
            if column not in {"代码", "名称"}
        )
        cards.append(
            f"""
            <details class="stock-card">
                <summary class="stock-card-title">
                    <strong>{name}</strong><span>{code}</span>
                </summary>
                <div class="stock-card-grid">{card_fields}</div>
            </details>
            """
        )

    st.markdown(
        f"""
        <div class="stock-desktop-table">
            <table><thead><tr>{headers}</tr></thead><tbody>{''.join(rows)}</tbody></table>
        </div>
        <div class="stock-mobile-cards">{''.join(cards)}</div>
        """,
        unsafe_allow_html=True,
    )


def build_excel(data: pd.DataFrame) -> bytes:
    """把真实筛选结果导出为 Excel 字节流。"""
    output = BytesIO()
    export_data = _display_data(data)
    export_data.to_excel(output, index=False, engine="openpyxl", sheet_name="筛选结果")
    return output.getvalue()


def build_strategy_report(
    final_top5: pd.DataFrame,
    initial_results: pd.DataFrame,
    excluded_results: pd.DataFrame,
    missing_records: pd.DataFrame,
    strategy_parameters: dict[str, object],
    updated_at: datetime,
) -> bytes:
    """生成包含五个审计工作表的策略报告。参数名保留用于兼容历史记录。"""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    output = BytesIO()
    top_columns = [
        "排名", "代码", "名称", "最新价", "涨跌幅", "量比", "换手率", "总市值",
        "主力资金净流入", "所属行业", "近5日板块强度", "VWAP状态", "尾盘最大回撤",
        "综合得分", "入选类型", "入选原因", "风险提示", "数据完整度",
        "当前行情数据源", "数据更新时间", "缺失字段", "评分依据",
    ]
    top_report = final_top5.reindex(columns=top_columns).copy()
    top_report.rename(
        columns={
            "综合得分": "综合评分",
            "资金表现得分": "资金分",
            "板块强度得分": "板块分",
            "技术形态得分": "技术分",
            "尾盘结构得分": "尾盘分",
            "市场情绪得分": "市场情绪分",
        },
        inplace=True,
    )
    parameter_frame = pd.DataFrame(
        [{"参数": key, "当前值": value} for key, value in strategy_parameters.items()]
    )
    sheets = {
        "最终Top5": top_report,
        "全部初筛结果": initial_results.copy(),
        "被排除股票": excluded_results.copy(),
        "数据缺失记录": missing_records.copy(),
        "策略参数快照": parameter_frame,
    }
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, frame in sheets.items():
            export = frame.copy()
            if export.empty and not len(export.columns):
                export = pd.DataFrame({"说明": ["本次扫描无记录"]})
            export = export.astype(object).where(pd.notna(export), None)
            export.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
            sheet = writer.book[sheet_name]
            column_count = max(1, len(export.columns))
            sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=column_count)
            sheet.cell(1, 1, f"A股尾盘策略报告 · {sheet_name}")
            sheet.cell(1, 1).font = Font(bold=True, color="FFFFFF", size=14)
            sheet.cell(1, 1).fill = PatternFill("solid", fgColor="102A43")
            sheet.cell(1, 1).alignment = Alignment(horizontal="left")
            for cell in sheet[3]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="176B87")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            sheet.freeze_panes = "A4"
            last_column = get_column_letter(column_count)
            sheet.auto_filter.ref = f"A3:{last_column}{sheet.max_row}"
            for index, column in enumerate(export.columns, start=1):
                values = [str(column)] + [str(value) for value in export[column].dropna().head(200)]
                width = min(42, max(10, max((len(value) for value in values), default=10) + 2))
                sheet.column_dimensions[get_column_letter(index)].width = width
                for cell in sheet[get_column_letter(index)][3:]:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                if column in {"数据完整度", "数据完整性", "高于VWAP占比", "尾盘最大回撤"}:
                    for cell in sheet[get_column_letter(index)][3:]:
                        cell.number_format = "0.0%"
            sheet.sheet_view.showGridLines = False
            sheet.cell(sheet.max_row + 2, 1, "本报告仅用于公开数据筛选和策略研究，不构成任何投资建议。")
            sheet.cell(sheet.max_row, 1).font = Font(color="C00000", italic=True)
        writer.book.properties.created = updated_at.replace(tzinfo=None)
    return output.getvalue()
